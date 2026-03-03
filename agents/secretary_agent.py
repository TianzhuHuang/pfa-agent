"""
私人秘书 (The Secretary)

职责：
  1. 编排 Scout → Analyst → Auditor 流水线
  2. 管理持仓配置（增删改 + CSV/JSON 批量导入）
  3. 管理数据源配置（RSS / Twitter / 监控 URL）
  4. 通过 Streamlit 提供 Web 控制中心

所有配置修改通过 Secretary 统一处理，保证数据一致性。
"""

from __future__ import annotations

import csv
import io
import json
import subprocess
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from agents.protocol import AgentMessage, make_request
from agents import scout_agent, analyst_agent, auditor_agent

CST = timezone(timedelta(hours=8))
PORTFOLIO_PATH = ROOT / "config" / "my-portfolio.json"
DATA_SOURCES_PATH = ROOT / "config" / "data-sources.json"

def _log(role: str, action: str, detail: str = ""):
    ts = datetime.now(CST).strftime("%H:%M:%S")
    print(f"[{ts}] {role:>10s} | {action} {detail}", flush=True)


# ===================================================================
# Portfolio Management
# ===================================================================

# load_portfolio, save_portfolio 委托给 pfa.portfolio_store
def load_portfolio() -> dict:
    from pfa.portfolio_store import load_portfolio as _load
    return _load()

def save_portfolio(data: dict) -> None:
    from pfa.portfolio_store import save_portfolio as _save
    _save(data)


_ACCOUNT_TYPE_MIGRATION = {"券商": "股票", "冷钱包": "数字货币", "实物": "其他"}


def _normalize_account(acc: Dict[str, Any]) -> Dict[str, Any]:
    """兼容旧数据：currency -> base_currency，account_type 旧值迁移，保留 balance"""
    out = dict(acc)
    if "base_currency" not in out and "currency" in out:
        out["base_currency"] = (out.get("currency") or "CNY").strip().upper()
    out.setdefault("base_currency", "CNY")
    at = out.get("account_type")
    if at in _ACCOUNT_TYPE_MIGRATION:
        out["account_type"] = _ACCOUNT_TYPE_MIGRATION[at]
    out.setdefault("account_type", "股票")
    out["balance"] = float(acc.get("balance", 0) or 0)
    return out


def get_accounts() -> List[Dict[str, Any]]:
    """Return structured accounts metadata list (含 base_currency, account_type)。"""
    portfolio = load_portfolio()
    raw = portfolio.get("accounts", [])
    return [_normalize_account(a) for a in raw]


def upsert_account(
    name: str,
    broker: str = "",
    currency: str = "CNY",
    account_type: str = "股票",
    balance: Optional[float] = None,
) -> Dict[str, Any]:
    """Create or update an account definition. 使用 name 作为人类可读主键。"""
    portfolio = load_portfolio()
    accounts = portfolio.setdefault("accounts", [])
    now = datetime.now(CST).isoformat()
    name = name.strip()
    broker = broker.strip()
    base_currency = (currency or "CNY").strip().upper()
    account_type = (account_type or "股票").strip() or "股票"

    for acc in accounts:
        if acc.get("name") == name:
            acc["broker"] = broker
            acc["base_currency"] = base_currency
            acc["account_type"] = account_type
            if balance is not None:
                acc["balance"] = float(balance)
            acc["updated_at"] = now
            if "currency" in acc:
                del acc["currency"]
            save_portfolio(portfolio)
            return _normalize_account(acc)

    acc = {
        "id": name,
        "name": name,
        "broker": broker,
        "base_currency": base_currency,
        "account_type": account_type,
        "balance": float(balance) if balance is not None else 0,
        "created_at": now,
        "updated_at": now,
    }
    accounts.append(acc)
    save_portfolio(portfolio)
    return acc


def update_account(account_id: str, **kwargs: Any) -> Optional[Dict[str, Any]]:
    """Update account by id. Supports name, base_currency, broker, account_type."""
    portfolio = load_portfolio()
    accounts = portfolio.setdefault("accounts", [])
    now = datetime.now(CST).isoformat()
    for acc in accounts:
        if acc.get("id") == account_id or acc.get("name") == account_id:
            if "name" in kwargs and kwargs["name"]:
                acc["name"] = str(kwargs["name"]).strip()
            if "base_currency" in kwargs and kwargs["base_currency"]:
                acc["base_currency"] = str(kwargs["base_currency"]).strip().upper()
            if "broker" in kwargs:
                acc["broker"] = str(kwargs["broker"] or "").strip()
            if "account_type" in kwargs and kwargs["account_type"]:
                acc["account_type"] = str(kwargs["account_type"]).strip()
            if "balance" in kwargs:
                acc["balance"] = float(kwargs["balance"] or 0)
            acc["updated_at"] = now
            if "currency" in acc:
                del acc["currency"]
            save_portfolio(portfolio)
            return _normalize_account(acc)
    return None


def delete_account(account_id: str) -> bool:
    """删除账户：从 accounts 移除，该账户下持仓迁移至「默认」。不允许删除「默认」账户。"""
    account_id = (account_id or "").strip()
    if not account_id or account_id == "默认":
        return False
    portfolio = load_portfolio()
    accounts = portfolio.setdefault("accounts", [])
    holdings = portfolio.setdefault("holdings", [])
    found = False
    new_accounts = []
    for acc in accounts:
        if acc.get("id") == account_id or acc.get("name") == account_id:
            found = True
            continue
        new_accounts.append(acc)
    if not found and not any(
        str(h.get("account", "")).strip() == account_id for h in holdings
    ):
        return False
    portfolio["accounts"] = new_accounts
    for h in holdings:
        if str(h.get("account", "")).strip() == account_id:
            h["account"] = "默认"
    has_default = any(
        a.get("id") == "默认" or a.get("name") == "默认" for a in new_accounts
    )
    if not has_default:
        portfolio["accounts"].insert(
            0,
            {
                "id": "默认",
                "name": "默认",
                "base_currency": "CNY",
                "broker": "",
                "account_type": "股票",
                "created_at": datetime.now(CST).isoformat(),
                "updated_at": datetime.now(CST).isoformat(),
            },
        )
    save_portfolio(portfolio)
    return True


def validate_portfolio() -> Tuple[bool, str]:
    result = subprocess.run(
        [sys.executable, str(ROOT / "scripts" / "validate_portfolio.py"),
         str(PORTFOLIO_PATH)],
        capture_output=True, text=True, cwd=str(ROOT),
    )
    return result.returncode == 0, result.stdout + result.stderr


def add_holding(symbol: str, name: str, market: str,
                cost_price: float = 0, quantity: int = 0,
                position_pct: float = 0, source: str = "manual") -> dict:
    """Add a single holding and return the updated portfolio."""
    portfolio = load_portfolio()
    now = datetime.now(CST).isoformat()
    h: Dict[str, Any] = {
        "symbol": symbol.strip(), "name": name.strip(),
        "market": market, "source": source, "updated_at": now,
    }
    if cost_price > 0:
        h["cost_price"] = cost_price
    if quantity > 0:
        h["quantity"] = quantity
    if position_pct > 0:
        h["position_pct"] = position_pct
    portfolio.setdefault("holdings", []).append(h)
    save_portfolio(portfolio)
    return portfolio


def remove_holding(symbol: str) -> dict:
    portfolio = load_portfolio()
    portfolio["holdings"] = [
        h for h in portfolio.get("holdings", []) if h.get("symbol") != symbol
    ]
    save_portfolio(portfolio)
    return portfolio


# ===================================================================
# Investment Memo (投资备忘录)
# ===================================================================

def add_memo(symbol: str, action: str, text: str,
             quantity_change: float = 0, price: float = 0) -> bool:
    """Add a memo entry to a holding's memo_history."""
    portfolio = load_portfolio()
    now = datetime.now(CST).isoformat()
    entry = {"time": now, "action": action, "text": text}
    if quantity_change:
        entry["quantity_change"] = quantity_change
    if price:
        entry["price"] = price

    for h in portfolio.get("holdings", []):
        if h["symbol"] == symbol:
            h.setdefault("memo_history", []).insert(0, entry)
            save_portfolio(portfolio)
            return True
    return False


def get_memos(symbol: str) -> List[Dict]:
    """Get all memos for a symbol, newest first."""
    portfolio = load_portfolio()
    for h in portfolio.get("holdings", []):
        if h["symbol"] == symbol:
            return h.get("memo_history", [])
    return []


def get_all_memos() -> List[Dict]:
    """Get all memos across all holdings, newest first."""
    portfolio = load_portfolio()
    all_memos = []
    for h in portfolio.get("holdings", []):
        sym = h["symbol"]
        name = h.get("name", sym)
        for m in h.get("memo_history", []):
            all_memos.append({**m, "symbol": sym, "name": name})
    all_memos.sort(key=lambda x: x.get("time", ""), reverse=True)
    return all_memos


def update_holdings_bulk(holdings: List[Dict]) -> dict:
    """Replace all holdings with a new list. Preserves currency, exchange."""
    portfolio = load_portfolio()
    now = datetime.now(CST).isoformat()
    clean = []
    for h in holdings:
        sym = str(h.get("symbol", "")).strip()
        if not sym:
            continue
        entry: Dict[str, Any] = {
            "symbol": sym,
            "name": str(h.get("name", "")).strip(),
            "market": h.get("market", "A"),
            "source": h.get("source", "manual"),
            "updated_at": now,
        }
        for num_key in ("cost_price", "quantity", "position_pct"):
            val = h.get(num_key)
            if val is not None and str(val).strip() != "":
                try:
                    fv = float(val)
                    if num_key == "quantity" or fv > 0:
                        entry[num_key] = fv
                except (ValueError, TypeError):
                    pass
        if h.get("account") is not None:
            entry["account"] = str(h.get("account", "")).strip() or "默认"
        if h.get("currency") in ("CNY", "USD", "HKD"):
            entry["currency"] = h.get("currency")
        if h.get("exchange"):
            entry["exchange"] = str(h.get("exchange", "")).strip()
        clean.append(entry)
    portfolio["holdings"] = clean
    save_portfolio(portfolio)
    return portfolio


def update_holding(symbol: str, account: str, new_symbol: Optional[str] = None, **kwargs: Any) -> Optional[Dict[str, Any]]:
    """Update a single holding by symbol+account. Returns updated holding or None.
    When new_symbol is provided and differs from symbol, replaces the holding (remove old, add new).
    """
    from pfa.symbol_utils import normalize_hk_symbol
    portfolio = load_portfolio()
    holdings = portfolio.get("holdings", [])
    now = datetime.now(CST).isoformat()
    acc = account.strip() or "默认"

    for i, h in enumerate(holdings):
        if str(h.get("symbol", "")).strip() != symbol.strip() or str(h.get("account", "")).strip() != acc:
            continue

        if new_symbol and new_symbol.strip() and new_symbol.strip() != symbol.strip():
            # Replace: remove old, add new with new_symbol
            base = dict(h)
            base.pop("symbol", None)
            base.pop("updated_at", None)
            new_sym = new_symbol.strip()
            if base.get("market") == "HK":
                new_sym = normalize_hk_symbol(new_sym)
            else:
                for suffix in (".HK", ".hk", ".SH", ".sh", ".SZ", ".sz"):
                    if new_sym.upper().endswith(suffix.upper()):
                        new_sym = new_sym[: -len(suffix)]
                        break
            updated_h = {
                "symbol": new_sym,
                "account": acc,
                "updated_at": now,
                **base,
            }
            if base.get("source") == "ocr":
                updated_h["ocr_confirmed"] = True
            if kwargs.get("quantity") is not None:
                try:
                    updated_h["quantity"] = float(kwargs["quantity"])
                except (ValueError, TypeError):
                    pass
            if kwargs.get("cost_price") is not None:
                try:
                    updated_h["cost_price"] = float(kwargs["cost_price"])
                except (ValueError, TypeError):
                    pass
            if kwargs.get("currency") in ("CNY", "USD", "HKD"):
                updated_h["currency"] = kwargs["currency"]
            if kwargs.get("market"):
                updated_h["market"] = str(kwargs["market"])[:2] or "A"
            if "exchange" in kwargs:
                updated_h["exchange"] = str(kwargs["exchange"]).strip() if kwargs["exchange"] else None
            if kwargs.get("name") is not None:
                updated_h["name"] = str(kwargs["name"]).strip()
            holdings[i] = updated_h
            save_portfolio(portfolio)
            return updated_h
        else:
            # In-place update
            if "quantity" in kwargs and kwargs["quantity"] is not None:
                try:
                    h["quantity"] = float(kwargs["quantity"])
                except (ValueError, TypeError):
                    pass
            if "cost_price" in kwargs and kwargs["cost_price"] is not None:
                try:
                    h["cost_price"] = float(kwargs["cost_price"])
                except (ValueError, TypeError):
                    pass
            if "currency" in kwargs and kwargs["currency"] in ("CNY", "USD", "HKD"):
                h["currency"] = kwargs["currency"]
            elif "currency" in kwargs and kwargs["currency"] is None:
                h.pop("currency", None)
            if "market" in kwargs and kwargs["market"]:
                h["market"] = str(kwargs["market"])[:2] or "A"
            if "exchange" in kwargs:
                if kwargs["exchange"]:
                    h["exchange"] = str(kwargs["exchange"]).strip()
                else:
                    h.pop("exchange", None)
            if "name" in kwargs and kwargs["name"] is not None:
                h["name"] = str(kwargs["name"]).strip()
            h["updated_at"] = now
            if h.get("source") == "ocr":
                h["ocr_confirmed"] = True
            save_portfolio(portfolio)
            return h
    return None


def parse_csv_holdings(csv_text: str) -> List[Dict]:
    """Parse CSV text into holdings list. Expects headers: symbol,name,market,cost_price,quantity"""
    reader = csv.DictReader(io.StringIO(csv_text))
    holdings = []
    for row in reader:
        h: Dict[str, Any] = {
            "symbol": row.get("symbol", row.get("代码", "")).strip(),
            "name": row.get("name", row.get("名称", "")).strip(),
            "market": row.get("market", row.get("市场", "A")).strip(),
            "source": "csv",
        }
        for key, cn in [("cost_price", "成本价"), ("quantity", "数量"),
                        ("position_pct", "仓位")]:
            val = row.get(key, row.get(cn, ""))
            if val:
                try:
                    h[key] = float(val)
                except ValueError:
                    pass
        if h["symbol"]:
            holdings.append(h)
    return holdings


def parse_json_holdings(json_text: str) -> List[Dict]:
    """Parse JSON text into holdings list. Accepts array or {holdings:[...]}"""
    data = json.loads(json_text)
    if isinstance(data, list):
        return data
    return data.get("holdings", [])


def parse_excel_holdings(excel_bytes: bytes) -> List[Dict]:
    """Parse Excel (.xlsx) into holdings list. Expects first row as header: symbol/代码, name/名称, market/市场, etc."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return []
    if not excel_bytes or len(excel_bytes) < 100:
        return []
    wb = load_workbook(io.BytesIO(excel_bytes), read_only=True, data_only=True)
    ws = wb.active
    if not ws:
        return []
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    header = [str(c or "").strip().lower() for c in rows[0]]
    col_map = {}
    for i, h in enumerate(header):
        if h in ("symbol", "代码", "code"):
            col_map["symbol"] = i
        elif h in ("name", "名称"):
            col_map["name"] = i
        elif h in ("market", "市场"):
            col_map["market"] = i
        elif h in ("cost_price", "成本价", "成本"):
            col_map["cost_price"] = i
        elif h in ("quantity", "数量"):
            col_map["quantity"] = i
        elif h in ("position_pct", "仓位"):
            col_map["position_pct"] = i
    if "symbol" not in col_map:
        for i, h in enumerate(header):
            if "symbol" in h or "代码" in h or "code" in h:
                col_map["symbol"] = i
                break
    if "symbol" not in col_map:
        return []
    holdings = []
    for row in rows[1:]:
        if col_map["symbol"] >= len(row):
            continue
        sym = str(row[col_map["symbol"]] or "").strip()
        if not sym:
            continue
        name_idx = col_map.get("name", col_map["symbol"])
        market_idx = col_map.get("market", col_map["symbol"])
        h: Dict[str, Any] = {
            "symbol": sym,
            "name": str(row[name_idx] or "").strip() if name_idx < len(row) else "",
            "market": (str(row[market_idx] or "A").strip() or "A") if market_idx < len(row) else "A",
            "source": "excel",
        }
        for key in ("cost_price", "quantity", "position_pct"):
            idx = col_map.get(key)
            if idx is not None and idx < len(row) and row[idx] is not None:
                try:
                    h[key] = float(row[idx])
                except (ValueError, TypeError):
                    pass
        holdings.append(h)
    return holdings


# ===================================================================
# Data Source Management
# ===================================================================

def _default_sources() -> dict:
    return {
        "rss_urls": [],
        "twitter_handles": [],
        "monitor_urls": [],
        "xueqiu_user_ids": [],
    }


def load_data_sources() -> dict:
    """Load from data-sources.json（本地）或 user_settings.data_sources（Supabase），按 user_id 隔离。"""
    try:
        from backend.database.supabase_store import use_supabase, get_data_sources
        from backend.context import get_current_user_id
        if use_supabase():
            return get_data_sources(get_current_user_id())
    except (ImportError, ValueError):
        pass
    if DATA_SOURCES_PATH.exists():
        with open(DATA_SOURCES_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    portfolio = load_portfolio()
    channels = portfolio.get("channels", {})
    sources = _default_sources()
    sources["rss_urls"] = channels.get("rss_urls", [])
    sources["twitter_handles"] = channels.get("twitter_handles", [])
    sources["xueqiu_user_ids"] = channels.get("xueqiu_user_ids", [])
    return sources


def save_data_sources(sources: dict) -> None:
    """保存到 data-sources.json（本地）或 user_settings.data_sources（Supabase）。"""
    try:
        from backend.database.supabase_store import use_supabase, save_data_sources as _save_sb
        from backend.context import get_current_user_id
        if use_supabase():
            _save_sb(sources, get_current_user_id())
            return
    except (ImportError, ValueError):
        pass
    with open(DATA_SOURCES_PATH, "w", encoding="utf-8") as f:
        json.dump(sources, f, ensure_ascii=False, indent=2)
    portfolio = load_portfolio()
    portfolio.setdefault("channels", {})
    portfolio["channels"]["rss_urls"] = [
        _extract_url(r) for r in sources.get("rss_urls", [])
    ]
    portfolio["channels"]["twitter_handles"] = sources.get("twitter_handles", [])
    portfolio["channels"]["xueqiu_user_ids"] = sources.get("xueqiu_user_ids", [])
    save_portfolio(portfolio)


def _extract_url(item) -> str:
    """Extract URL string from an RSS item (could be str or dict)."""
    if isinstance(item, str):
        return item
    return item.get("url", "")


def add_rss_source(url: str, name: str = "", category: str = "") -> dict:
    """Add an RSS source (structured object format)."""
    sources = load_data_sources()
    existing_urls = {_extract_url(r) for r in sources.get("rss_urls", [])}
    url = url.strip()
    if url and url not in existing_urls:
        entry: Dict[str, Any] = {"url": url, "enabled": True}
        if name:
            entry["name"] = name
        if category:
            entry["category"] = category
        sources.setdefault("rss_urls", []).append(entry)
    save_data_sources(sources)
    return sources


def remove_rss_source(url: str) -> dict:
    """Remove an RSS source by URL."""
    sources = load_data_sources()
    sources["rss_urls"] = [
        r for r in sources.get("rss_urls", [])
        if _extract_url(r) != url
    ]
    save_data_sources(sources)
    return sources


def update_rss_source(old_url: str, url: str, name: str = "", category: str = "") -> dict:
    """Update an RSS source by old URL."""
    sources = load_data_sources()
    rss_list = sources.get("rss_urls", [])
    for i, r in enumerate(rss_list):
        if _extract_url(r) == old_url:
            entry: Dict[str, Any] = {"url": url.strip(), "enabled": r.get("enabled", True)}
            if name:
                entry["name"] = name.strip()
            if category:
                entry["category"] = category.strip()
            if r.get("via"):
                entry["via"] = r["via"]
            rss_list[i] = entry
            break
    save_data_sources(sources)
    return sources


def toggle_api_source(source_type: str, enabled: bool) -> dict:
    """Toggle API source enabled/disabled by type."""
    sources = load_data_sources()
    api_list = sources.get("api_sources", [])
    for src in api_list:
        if src.get("type") == source_type:
            src["enabled"] = enabled
            break
    save_data_sources(sources)
    return sources


def set_rsshub_url(url: str) -> dict:
    """Set self_hosted_rsshub URL."""
    sources = load_data_sources()
    sources["self_hosted_rsshub"] = (url or "").strip()
    save_data_sources(sources)
    return sources


def add_source(source_type: str, value: str) -> dict:
    """Add a simple string-type source (twitter, xueqiu, monitor_urls)."""
    sources = load_data_sources()
    key = source_type if source_type in sources else f"{source_type}s"
    if key not in sources:
        sources[key] = []
    value = value.strip()
    if value and value not in sources[key]:
        sources[key].append(value)
    save_data_sources(sources)
    return sources


def remove_source(source_type: str, value: str) -> dict:
    sources = load_data_sources()
    key = source_type if source_type in sources else f"{source_type}s"
    if key in sources:
        sources[key] = [v for v in sources[key] if v != value]
    save_data_sources(sources)
    return sources


def set_xueqiu_user_ids(ids: list) -> dict:
    """Replace xueqiu_user_ids with the given list (for sync follow list)."""
    sources = load_data_sources()
    sources["xueqiu_user_ids"] = [str(i).strip() for i in (ids or []) if str(i).strip()]
    save_data_sources(sources)
    return sources


# ===================================================================
# Pipeline Orchestration
# ===================================================================

def run_full_pipeline(
    holdings: List[Dict],
    hours: int = 72,
    do_audit: bool = True,
) -> Dict[str, Any]:
    """Run Scout → Analyst → Auditor pipeline."""
    result: Dict[str, Any] = {
        "status": "ok", "scout_result": None,
        "analyst_result": None, "auditor_result": None,
        "pipeline_log": [],
    }

    def log(msg: str):
        result["pipeline_log"].append(msg)
        _log("Secretary", msg)

    log("Scout → 开始抓取新闻...")
    scout_req = make_request("secretary", "scout", "fetch_all",
                             holdings=holdings, hours=hours)
    scout_resp = scout_agent.handle(scout_req)

    if scout_resp.status != "ok":
        log(f"Scout 失败: {scout_resp.error}")
        result["status"] = "error"
        result["error"] = scout_resp.error
        return result

    total = scout_resp.payload.get("total", 0)
    log(f"Scout ← 完成，共 {total} 条新闻")
    result["scout_result"] = scout_resp.payload

    if total == 0:
        log("无新闻，流程结束")
        return result

    log("Analyst → 开始深度分析...")
    analyst_req = make_request("secretary", "analyst", "analyze",
                               items=scout_resp.payload["items"],
                               holdings=holdings)
    analyst_resp = analyst_agent.handle(analyst_req)

    if analyst_resp.status != "ok":
        log(f"Analyst 失败: {analyst_resp.error}")
        result["status"] = "partial"
        result["error"] = analyst_resp.error
        return result

    log(f"Analyst ← 完成 (model={analyst_resp.payload.get('model', '?')})")
    result["analyst_result"] = analyst_resp.payload

    if do_audit:
        log("Auditor → 开始交叉验证...")
        auditor_req = make_request("secretary", "auditor", "audit",
                                    analysis=analyst_resp.payload.get("analysis", ""),
                                    items=scout_resp.payload["items"],
                                    holdings=holdings)
        auditor_resp = auditor_agent.handle(auditor_req)

        if auditor_resp.status != "ok":
            log(f"Auditor 失败: {auditor_resp.error}（分析仍有效，未经审核）")
            result["status"] = "partial"
        else:
            log(f"Auditor ← 完成 (backend={auditor_resp.payload.get('backend', '?')})")
            result["auditor_result"] = auditor_resp.payload
    else:
        log("Auditor 跳过")

    log("流程完成")
    return result


def run_full_pipeline_streaming(holdings: List[Dict], hours: int = 72, do_audit: bool = False):
    """Generator: yields status events then final result. For SSE streaming."""
    result: Dict[str, Any] = {
        "status": "ok", "scout_result": None,
        "analyst_result": None, "auditor_result": None,
        "pipeline_log": [],
    }

    def log(msg: str):
        result["pipeline_log"].append(msg)
        _log("Secretary", msg)

    import time
    try:
        log("Scout → 开始抓取新闻...")
        yield {"type": "status", "message": "正在巡视全球财经情报，搜索相关新闻...", "stage": "scout_start", "color": "blue"}

        scout_req = make_request("secretary", "scout", "fetch_all",
                                 holdings=holdings, hours=hours)
        scout_resp = scout_agent.handle(scout_req)

        if scout_resp.status != "ok":
            log(f"Scout 失败: {scout_resp.error}")
            result["status"] = "error"
            result["error"] = scout_resp.error
            yield {"type": "error", "error": scout_resp.error}
            return

        total = scout_resp.payload.get("total", 0)
        log(f"Scout ← 完成，共 {total} 条新闻")
        yield {"type": "status", "message": f"已获取 {total} 条最新资讯，正在筛选高价值信息...", "stage": "scout_done", "color": "blue"}

        # 实时透传抓取到的新闻标题（Manus 风格 Ticker）
        raw_items = scout_resp.payload.get("items", [])
        for i, it in enumerate(raw_items[:30]):  # 最多推送 30 条，避免刷屏
            title = (it.get("title") or it.get("content_snippet") or "")[:60]
            if title:
                yield {"type": "ticker", "message": f"[Scout] 发现: {title}...", "color": "blue", "source": "scout"}

        result["scout_result"] = scout_resp.payload

        if total == 0:
            log("无新闻，流程结束")
            yield {"type": "done", "result": result}
            return

        log("Analyst → 开始深度分析...")
        yield {"type": "status", "message": "正在根据你的持仓风险模型进行深度推演...", "stage": "analyst_start", "color": "purple"}

        # 后台线程调用 Analyst，主线程每 2 秒 yield 心跳，避免 Next.js 代理 30s 超时断开
        import threading
        analyst_resp = [None]
        analyst_err = [None]

        def _run_analyst():
            try:
                raw_items = scout_resp.payload.get("items", [])
                if isinstance(raw_items, list):
                    # 雪球大 V 内容优先排序
                    sorted_items = sorted(
                        raw_items,
                        key=lambda x: (0 if (x.get("source") or "").startswith("xueqiu") else 1, -(len(x.get("content_snippet") or ""))),
                    )
                    items_to_analyze = sorted_items[:10]
                else:
                    items_to_analyze = []
                log(f"Analyst 调用 LLM，新闻数={len(items_to_analyze)}")
                analyst_req = make_request("secretary", "analyst", "analyze",
                                           items=items_to_analyze,
                                           holdings=holdings)
                analyst_resp[0] = analyst_agent.handle(analyst_req)
            except Exception as e:
                log(f"Analyst 异常: {e}")
                analyst_err[0] = e

        t = threading.Thread(target=_run_analyst, daemon=True)
        t.start()
        # 具体分析维度描述，替代死板心跳
        analyst_steps = [
            "正在比对持仓标的与新闻关联...",
            "正在检索 A 股电力板块情绪...",
            "正在分析蔚来/新能源财报数据...",
            "正在筛选必读要点...",
            "正在生成异动标的建议...",
            "正在计算市场情绪得分...",
            "正在整理报告排版...",
        ]
        hb_idx = 0
        yield {"type": "status", "message": analyst_steps[0], "stage": "analyst_heartbeat", "color": "purple"}
        log("Analyst 启动")
        hb_idx = 1
        while t.is_alive():
            time.sleep(2)
            if t.is_alive():
                msg = analyst_steps[hb_idx % len(analyst_steps)]
                yield {"type": "status", "message": msg, "stage": "analyst_heartbeat", "color": "purple"}
                # 同时推送 Ticker 风格的分析碎片，增强颗粒感
                yield {"type": "ticker", "message": f"[Analyst] {msg}", "color": "purple", "source": "analyst"}
                log(f"Analyst 进度 ({hb_idx})")
                hb_idx += 1
        t.join(timeout=1)
        if analyst_err[0]:
            yield {"type": "error", "error": f"分析失败: {str(analyst_err[0])[:200]}"}
            return
        analyst_resp = analyst_resp[0]

        if analyst_resp.status != "ok":
            log(f"Analyst 失败: {analyst_resp.error}")
            result["status"] = "partial"
            result["error"] = analyst_resp.error
            result["analyst_result"] = analyst_resp.payload
            yield {"type": "status", "message": f"分析异常: {analyst_resp.error}", "stage": "analyst_fail", "color": "purple"}
        else:
            log(f"Analyst ← 完成 (model={analyst_resp.payload.get('model', '?')})")
            yield {"type": "status", "message": "深度分析报告已生成，正在整理排版...", "stage": "analyst_done", "color": "purple"}

        result["analyst_result"] = analyst_resp.payload

        if do_audit:
            log("Auditor → 开始交叉验证...")
            yield {"type": "status", "message": "交叉验证中...", "stage": "auditor_start", "color": "purple"}
            auditor_req = make_request("secretary", "auditor", "audit",
                                       analysis=analyst_resp.payload.get("analysis", ""),
                                       items=scout_resp.payload["items"],
                                       holdings=holdings)
            auditor_resp = auditor_agent.handle(auditor_req)
            if auditor_resp.status != "ok":
                log(f"Auditor 失败: {auditor_resp.error}（分析仍有效，未经审核）")
                result["status"] = "partial"
                yield {"type": "status", "message": "审核跳过（分析仍有效）", "stage": "auditor_fail", "color": "purple"}
            else:
                log(f"Auditor ← 完成 (backend={auditor_resp.payload.get('backend', '?')})")
                yield {"type": "status", "message": "交叉验证完成", "stage": "auditor_done", "color": "purple"}
            result["auditor_result"] = auditor_resp.payload
        else:
            log("Auditor 跳过")
            yield {"type": "status", "message": "审核跳过（分析仍有效）", "stage": "auditor_skip", "color": "gray"}

        log("流程完成")
        yield {"type": "status", "message": "晨报已就绪", "stage": "done", "color": "green"}
        yield {"type": "done", "result": result}
    except Exception as e:
        result["status"] = "error"
        result["error"] = str(e)
        yield {"type": "error", "error": str(e)}


def fetch_single_symbol(
    symbol: str, name: str, market: str, hours: int = 72
) -> Dict[str, Any]:
    req = make_request("secretary", "scout", "fetch_news",
                       symbol=symbol, name=name, market=market, hours=hours)
    resp = scout_agent.handle(req)
    if resp.status != "ok":
        return {"status": "error", "error": resp.error}
    return {"status": "ok", **resp.payload}
